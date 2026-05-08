"""Tiny helpers to render rows of dicts as a styled XLSX bytes stream."""
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
ALT_FILL = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")


def _coerce(value):
    """Make a cell value openpyxl-friendly."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        # Strip tz so Excel renders as a plain date-time
        return value.replace(tzinfo=None)
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return value


def rows_to_xlsx(*, sheet_title: str, columns: list[tuple[str, str]], rows: list[dict]) -> bytes:
    """Render a list of dicts as styled XLSX bytes.

    `columns` is a list of (header_label, dict_key) tuples.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel sheet name limit

    # Header row
    for col_idx, (label, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (_, key) in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_coerce(row.get(key)))
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    # Column widths
    for col_idx, (label, key) in enumerate(columns, start=1):
        max_len = len(str(label))
        for row in rows:
            v = row.get(key)
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
